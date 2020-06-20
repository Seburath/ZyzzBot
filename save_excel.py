from openpyxl import load_workbook


def save(genere, height, weight, neck, waist, hip, body_fat, ffmi, body_fat_100, ffmi_100, index_100):
    data = load_workbook(filename = 'data.xlsx')
    sheet = data["Hoja1"]

    def read(sheet, col, row):
        return sheet.cell(column=col, row=row).value

    def write(sheet, col, row, val):
        sheet.cell(column=col, row=row).value = val
        return

    y = sheet.max_row + 1

    write(sheet, 1, y, genere)
    write(sheet, 2, y, height)
    write(sheet, 3, y, weight)
    write(sheet, 4, y, neck)
    write(sheet, 5, y, waist)
    write(sheet, 6, y, hip)
    write(sheet, 7, y, str(body_fat))
    write(sheet, 8, y, str(ffmi))
    write(sheet, 9, y, str(int(body_fat_100)))
    write(sheet, 10, y, str(int(ffmi_100)))
    write(sheet, 11, y, str(int(index_100)))

    print('registro numero: ' + str(sheet.max_row) + ' añadido')
    data.save('data.xlsx')

    return
